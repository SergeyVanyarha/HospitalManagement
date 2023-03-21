from datetime import datetime
from odoo import models, fields
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json
import io
import base64


class HospitalVisitReport(models.TransientModel):
    _name = 'hospital.visit.report'
    _description = 'Hospital Visit Report'

    start_date = fields.Datetime(
        string='Start Date',
    )

    end_date = fields.Datetime(
        string='End Date',
    )

    patient_id = fields.Many2one(
        'hospital.patient',
        string='Patient',
    )

    doctor_id = fields.Many2one(
        'hospital.doctor',
        string='Doctor',
    )

    file_data = fields.Binary(
        string='File',
    )

    filename = fields.Char(
        string='File Name',
    )

    is_finished = fields.Boolean(
        string='Is Finished',
        default=True
    )

    def open_visit_action(self):
        domain = []
        if self.patient_id:
            domain.append(('patient_id', '=', self.patient_id.id))
        if self.start_date:
            domain.append(('visit_datetime', '>=', self.start_date))
        if self.end_date:
            domain.append(('visit_datetime', '<=', self.end_date))
        if self.doctor_id:
            domain.append(('doctor_id', '=', self.doctor_id.id))
        if self.is_finished:
            domain.append(('visit_datetime', '<=', fields.Datetime.now()))
        return {
            'type': 'ir.actions.act_window',
            'name': 'Patient Visits',
            'res_model': 'hospital.patient.visit',
            'view_mode': 'tree,form,calendar',
            'domain': domain,
        }

    def export_excel_report(self):
        book = Workbook()
        sheet = book.active

        # Fill headers
        header = (
            'Phone', 'Doctor', 'Visit Datetime', 'Diagnosis',
            'Treatment', 'Notes'
        )
        sheet.append(header)
        for i in list(range(1, len(header) + 1)):
            sheet.column_dimensions[get_column_letter(i)].width = 20

        # Find visits and fill them into the sheet
        domain = []
        if self.patient_id:
            domain.append(('patient_id', '=', self.patient_id.id))
        if self.start_date:
            domain.append(('visit_datetime', '>=', self.start_date))
        if self.end_date:
            domain.append(('visit_datetime', '<=', self.end_date))
        if self.doctor_id:
            domain.append(('doctor_id', '=', self.doctor_id.id))
        if self.is_finished:
            domain.append(('visit_datetime', '<=', fields.Datetime.now()))
        visits = self.env['hospital.patient.visit'].search(domain)
        for visit in visits:
            sheet.append((
                visit.patient_id.phone, visit.doctor_id.name, str(visit.visit_datetime), visit.diagnosis or "",
                visit.treatment or "", visit.notes or ""
            ))

        # Save the file
        fp = io.BytesIO()
        book.save(fp)
        self.file_data = base64.encodebytes(fp.getvalue())
        self.filename = 'export_hospital_visit_report %s.xlsx' % datetime.strftime(
            datetime.now(), '%d-%m-%Y %H:%M:%S')
        fp.close()
        return {
            'name': 'Download excel Hospital Visit Report',
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=hospital.visit.report&id={self.id}&field=file_data'
                   '&filename_field=filename&download=true',
            'target': 'self',
        }

